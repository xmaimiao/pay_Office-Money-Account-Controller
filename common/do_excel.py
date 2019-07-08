from openpyxl import load_workbook
from common import contants

class Cases:
    def __init__(self):
        self.id=None
        self.title=None
        self.url=None
        self.method=None
        self.headers=None
        self.json=None
        self.expect=None


class Doexcel:
    def __init__(self,filename,sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb[self.sheet_name]

    def read_excel(self):
        case = []
        for row in range(2,self.sheet.max_row+1):
            cases = Cases()
            cases.id = self.sheet.cell(row,1).value
            cases.title = self.sheet.cell(row,2).value
            cases.url = self.sheet.cell(row,3).value
            cases.method = self.sheet.cell(row,4).value
            cases.headers = self.sheet.cell(row,5).value
            cases.json = self.sheet.cell(row,6).value
            cases.data = self.sheet.cell(row,7).value
            cases.expect = self.sheet.cell(row,8).value
            case.append(cases)
        return case

    def write_excel(self,row,actual,result,token_msg=0):
        self.sheet.cell(row,9).value = actual
        self.sheet.cell(row,10).value = result
        self.sheet.cell(row,11).value = token_msg
        self.wb.save(self.filename)
        self.wb.close()

if __name__ == '__main__':
    do = Doexcel(contants.case_dir, 'getgroupid ')
    do.read_excel()
    do.write_excel(2,2,3,4)
    print(do)
